# -*- coding: utf-8 -*-
"""
Created Feb. 2018 @author:Sean T. Smith
"""
from numpy import empty, sqrt
from scipy.stats import t
from pandas import DataFrame, Series, ExcelFile, isnull
from openpyxl import load_workbook

# from pyregress import *


class InputParser:
    def __init__(self, input_file):
        """
        Create a class to import and parse the excel spreadsheet that is used
        as an input file for V/UQ-predictivity.
        """
        self.file_name = input_file
        # Import the excel file:
        self.xlfile = ExcelFile(self.file_name)  # to retrieve & work w/ input

    def parse_expt(self):
        """
        Parse (w/ initial processing) of the excel worksheet for the
        experimental data. The worksheet must have the following format:
        -Sheet titled "Experiments",
        -First row ignored - use for notes,
        -Rows 3-7 ignored - used to document format,
        -First column labels the QOI type or group,
        -Next column labels the specific identifier or scenario for each QOI,
        -Column titled 'Outlier Flag' with values left blank unless the
         type/scenario combination is to be ignored,
        -Two output columns: 'Data μ' (midpoint of the data), and
                             'Data σ' (one-sided error of the data),
        -Column titled 'Input Mode' indicates which method is used to specify
         the acceptable data range (empty fills from above),
         Options:              Data Values:
         -'samples'          -data samples/replicates (multiple values),
         -'stats'            -number of samples, sample mean, sample std.,
         -'prior+samples'    -prior σ, prior equiv. No., samples,
         -'prior+stats'      -prior σ, prior equiv. No., stats,
         -'mid.-err.-conf.'  -data mid., error (1 sided), confidence level, No.
        -Remaining columns titled 'Data Values' specify data,  statistics, etc.
         in the format corresponding to 'Input Mode'.
        """
        # Parse the experimental data and calculate bounds:
        expt = self.xlfile.parse('Experiments', header=1, index_col=[0, 1],
                                 skiprows=[2, 3, 4, 5, 6])
        # Satisfy the fill-from-above expectations:
        expt['Input Mode'].fillna(method='pad', inplace=True)
        prior_σ, prior_n, confint, error = None, None, None, None
        for ind, e in expt.iterrows():
            if e['Input Mode'].startswith('prior'):
                if not isnull(e['Data Values':][0]):
                    prior_σ = e['Data Values':][0]
                else:
                    expt.at[e.name, e['Data Values':].index[0]] = prior_σ
                if not isnull(e['Data Values':][1]):
                    prior_n = e['Data Values':][1]
                else:
                    expt.at[e.name, e['Data Values':].index[1]] = prior_n
            if e['Input Mode'] == 'mid.-err.-conf.':
                if not isnull(e['Data Values':][1]):
                    error = e['Data Values':][1]
                else:
                    expt.at[e.name, e['Data Values':].index[1]] = error
                if not isnull(e['Data Values':][2]):
                    confint = e['Data Values':][2]
                else:
                    expt.at[e.name, e['Data Values':].index[2]] = confint
                if not isnull(e['Data Values':][3]):
                    prior_n = e['Data Values':][3]
                else:
                    expt.at[e.name, e['Data Values':].index[3]] = prior_n
        # Remove indicated outliers from the Experiments worksheet:
        inliers = self.inliers = isnull(expt['Outlier Flag'])
        iall = self.iall = [ia for ia in range(len(expt)) if inliers[ia]]
        Nqoi = self.Nqoi = len(iall)
        expt = expt[inliers]
        # Process the data according to the specified input mode:
        stats = DataFrame(index=expt.index, columns=['n0', 'n', 'm', 's'])
        for i in range(Nqoi):
            e = expt.iloc[i]
            mode = e['Input Mode']
            data = e['Data Values':]
            if mode == 'samples':
                n0, n, m, s = 0, data.count(), data.mean(), data.std(ddof=0)
            elif mode == 'stats':
                n0, n, m, s = 0, data[0], data[1], data[2]
            elif mode.startswith('prior'):
                s0, n0 = data[0], data[1]
                if mode.endswith('samples'):
                    n = data[2:].count()
                    m = data[2:].mean()
                    s = data[2:].std(ddof=0)
                elif mode.endswith('stats'):
                    n, m, s = data[2], data[3], data[4]
                # s0 = (n0 - 2) / n0 * Eσ0
                if n0 + n == 1:
                    s = 0
                else:
                    s = sqrt((n0 * s0**2 + n * s**2) / (n0 + n - 1))
            elif mode == 'mid.-err.-conf.':
                m, err, ci, n0, n = data[0], data[1], data[2], 0, data[3]
                s = err / (sqrt(1 + 1 / n) * t.ppf((ci + 1) / 2, n + 1))
            stats.at[e.name, 'n0'] = n0
            stats.at[e.name, 'n'] = n
            stats.at[e.name, 'm'] = m
            stats.at[e.name, 's'] = s
            stats = stats.astype('float64')

        return stats, Nqoi

    def parse_dsgn(self):
        """
        Parse the excel worksheet for the experimental design.
        The worksheet must have the following format:
        -Sheet titled "Parameters & Design",
        -First row has two sections 'Inputs (priors & design)' &
         'Outputs (posterior μ, σ and Σ)',
        -Second row provides the parameter labels  (repeated in each section),
        -First column has entries 'µ', 'σ' & case the IDs that will be
         used on the simulation & surrogate modeling sheets,
        -Remaining input columns provide the design values for the parameters.
        """
        # Parse the simulation design:
        dsgn = self.xlfile.parse('Parameters & Design', header=[0, 1])
        dsgn = dsgn['Inputs (priors & design)']
        param_priors = dsgn.loc[['µ', 'σ']]
        dsgn = dsgn.drop(['µ', 'σ'])
        Nmp = dsgn.shape[1]
        return dsgn, Nmp, param_priors

    def parse_sim(self):
        """
        Parse the excel worksheet for the simulation data.
        The worksheet must have the following format:
        -Sheet titled "Simulation Data",
        -First row ignored - use for notes,
        -First column labels the QOI type or group (matches experimental data),
        -Next column labels the specific identifier or scenario (matching
         experimental data),
        -Remaining columns are for the multiple simulation runs, with the
         corresponding case identifier from the simulation design sheet.
        Note: must parse expt beforehand to correctly identify inliers.
        """
        # Parse the simulation data:
        sims = self.xlfile.parse('Simulation Data', header=1, index_col=[0, 1])
        if hasattr(self, 'inliers'):
            sims = sims[self.inliers]
        return sims

    def parse_surr(self, dsgn, sims, rewrite_φ=True):
        """
        Parse the excel worksheet for the experimental design.
        The worksheet must have the following format:
        -Sheet titled "Surrogate Model",
        -First row ignored - use for notes.
        -First column labels the QOI type or group (matches experimental data),
        -Next column labels the specific identifier or scenario (matching
         experimental data),
        -Column titled 'Method' indicating the technique, options:
          'linear regression', 'quadratic regression', 'piecewise linear',
          and 'Gaussian process' (which requires the remaining fields);
        -Column titled 'Xscaling', scaling method for independent variables,
        -Column titled 'Explicit Basis', polynomial bases for interpolation,
        -Column titled 'Kernel', specifies the kernel with formatting place-
         holders (eg. {}) for any hyper-parameters (empty fills from above),
        -Column titled 'Optimize', indicate whether to optimize hyper-params.,
        -Columns titled 'Parameters' and all that follow specify the
         hyper-parameters or their initial guess - used in the order specified.
        """
        # Parse the surrogate model specification and build interpolants:
        spec = self.xlfile.parse('Surrogate Model', header=1, index_col=[0, 1])
        if rewrite_φ:
            worksheet = self.xlwb.get_sheet_by_name('Surrogate Model')
            for icol in range(1, worksheet.max_column + 1):
                index = worksheet.cell(row=2, column=icol).value
                if index == 'Parameters':
                    col_φ = icol
                    break
        # Satisfy the fill-from-above expectations:
        spec['Method'].fillna(method='pad', inplace=True)
        gp_slice = spec['Method'] == 'Gaussian process'
        tmp = spec.loc[gp_slice].fillna(method='pad')
        spec.update(tmp)
        spec = spec[self.inliers]
        # Create a pandas series for the surrogates and fill it:
        fsurr = Series(empty(self.Nqoi, dtype=object), index=sims.index)
        for i in range(self.Nqoi):
            e = spec.iloc[i]
            if e['Method'] == 'piecewise linear':
                fsurr.iloc[i] = PiecewiseLinear(dsgn.values, sims.values[i, :])
            elif e['Method'] == 'linear regression':
                # TODO: Add regression tool.
                pass
            elif e['Method'] == 'quaratic regression':
                # TODO: Add regression tool.
                pass
            elif e['Method'] == 'Gaussian process':
                φ = e['Parameters':].dropna()
                kernel = eval(e['Kernel'].format(*φ))
                xscale = e['Xscaling']
                bases = eval(e['Explicit Basis'])
                opt = e['Optimize']
                if opt == 'verbose' or opt == 'v':
                    print('Generating GP for QOI: {}, {}...'.format(*e.name))
                fsurr.iloc[i] = GPI(dsgn.values, sims.values[i, :], kernel,
                                    explicit_basis=bases, Xscaling=xscale,
                                    optimize=opt)
                # Test the interpolation:
                try:
                    fsurr.iloc[i].loo()
                except ValidationError as err:
                    print('QOI {:d} failing cross validation on {:d} '
                          'of {:d} points!'.format(i + 1, err.N3, err.Nd))
                # Rewrite optimized parameters back to the excel file:
                φ = fsurr.iloc[i].kernel.get_φ(trans=False)
                if opt and rewrite_φ:
                    row = self.iall[i] + 3
                    for iφ in range(len(φ)):
                        worksheet.cell(row=row, column=col_φ+iφ, value=φ[iφ])
        return fsurr

    def write_posteriors(self, μ, Σ, yd, σy):
        # Parameters:
        Np = μ.shape[0]
        self.xlwb = load_workbook(self.file_name)
        # worksheet = self.xlwb.get_sheet_by_name('Parameters & Design')
        worksheet = self.xlwb['Parameters & Design']

        # TODO: Write the parameter posterior.
        for j in range(Np):
            worksheet.cell(row=3, column=Np+2+j, value=μ[j])
            worksheet.cell(row=4, column=Np+2+j, value=sqrt(Σ[j, j]))
            for i in range(Np):
                worksheet.cell(row=5+i, column=Np+2+j, value=Σ[i, j])

        self.xlwb.save(self.file_name)

        # Expterimental measurements & error:
        self.xlwb = load_workbook(self.file_name)
        # worksheet = self.xlwb.get_sheet_by_name('Experiments')
        worksheet = self.xlwb['Experiments']
        for icol in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=2, column=icol).value == 'Data μ':
                col_μ = icol
                break
        for icol in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=2, column=icol).value == 'Data σ':
                col_σ = icol
                break

        for i in range(self.Nqoi):
            worksheet.cell(row=self.iall[i] + 8, column=col_μ, value=yd[i])
            worksheet.cell(row=self.iall[i] + 8, column=col_σ, value=σy[i])
        self.xlwb.save(self.file_name)
        return None


if __name__ == '__main__':
    file = 'BSF_year4_v2.xlsx'
    my_parser = InputParser(file)

    expt, Nqoi = my_parser.parse_expt()
    print('There are {} QOIs'.format(Nqoi))
    print(expt)
    print(' ')

    dsgn, Nmp, param_priors = my_parser.parse_dsgn()
    print('There are {} model parameters'.format(Nmp))
    print(dsgn)
    print(param_priors)
    print(' ')

    sims = my_parser.parse_sim()
    print(sims)
    print(' ')
